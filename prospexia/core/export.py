"""Export des prospects en CSV / Excel / PDF."""
from __future__ import annotations

import csv
import datetime as dt
import html
import re
from pathlib import Path

from prospexia.core.models import Prospect, Verdict


def _legal(p: Prospect, attr: str, default: str = "") -> str:
    if not p.legal:
        return default
    v = getattr(p.legal, attr, default)
    if isinstance(v, list):
        return " ; ".join(v)
    return str(v) if v is not None else default


COLUMNS = [
    ("Entreprise", lambda p: p.company.name),
    ("Secteur", lambda p: p.company.sector),
    ("Ville", lambda p: p.company.city),
    ("Adresse", lambda p: p.company.address),
    ("Téléphone", lambda p: p.company.phone),
    ("Site web", lambda p: p.company.website),
    ("État du site", lambda p: p.website.status.label),
    ("Score présence", lambda p: p.score),
    ("Verdict", lambda p: p.verdict.label),
    ("Opportunité", lambda p: p.opportunity),
    ("Réseaux sociaux", lambda p: ", ".join(p.presence.socials.values())),
    ("Résultats web", lambda p: p.presence.search_hits),
    ("Avis Google", lambda p: p.company.reviews_count),
    ("Note Google", lambda p: p.company.rating if p.company.rating is not None else ""),
    ("Diagnostic", lambda p: " · ".join(p.reasons)),
    # --- registre officiel ---
    ("Registre", lambda p: _legal(p, "registry")),
    ("Identifiant", lambda p: (f"{_legal(p, 'identifier_label')} {_legal(p, 'identifier')}").strip()),
    ("Identifiant secondaire", lambda p: (f"{_legal(p, 'secondary_label')} {_legal(p, 'secondary_id')}").strip()),
    ("Dénomination légale", lambda p: _legal(p, "legal_name")),
    ("Forme juridique", lambda p: _legal(p, "legal_form")),
    ("Code activité", lambda p: (f"{_legal(p, 'activity_code')} {_legal(p, 'activity_label')}").strip()),
    ("Date de création", lambda p: _legal(p, "creation_date")),
    ("Effectif", lambda p: _legal(p, "headcount")),
    ("État légal", lambda p: _legal(p, "status")),
    ("Siège", lambda p: _legal(p, "address")),
    ("Dirigeants", lambda p: _legal(p, "managers")),
    ("N° TVA", lambda p: _legal(p, "vat_number")),
    ("Fiche registre", lambda p: _legal(p, "source_url")),
    ("Appariement registre", lambda p: ("auto" if p.legal.matched else "à vérifier") if p.legal else ""),
    ("Source", lambda p: p.company.source),
    ("Lien carte", lambda p: p.company.maps_url),
]


def export_csv(prospects: list[Prospect], path: Path) -> Path:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f, delimiter=";")
        w.writerow([c for c, _ in COLUMNS])
        for p in prospects:
            w.writerow([fn(p) for _, fn in COLUMNS])
    return path


def export_xlsx(prospects: list[Prospect], path: Path) -> Path:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "Prospects"
    ws.append([c for c, _ in COLUMNS])
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="5B4BDB")
        cell.alignment = Alignment(vertical="center")
    fills = {"Prioritaire": "DCFCE7", "Cible": "FEF9C3", "Hors cible": "FEE2E2"}
    verdict_col = [c for c, _ in COLUMNS].index("Verdict") + 1
    for p in prospects:
        ws.append([fn(p) for _, fn in COLUMNS])
        color = fills.get(p.verdict.label)
        if color:
            ws.cell(row=ws.max_row, column=verdict_col).fill = PatternFill("solid", fgColor=color)
    for i in range(1, len(COLUMNS) + 1):
        width = max(12, min(50, max(len(str(ws.cell(row=r, column=i).value or ""))
                                    for r in range(1, ws.max_row + 1)) + 2))
        ws.column_dimensions[get_column_letter(i)].width = width
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    wb.save(path)
    return path


# --------------------------------------------------------------------------- PDF
_VERDICT_BG = {Verdict.PRIORITY: "#dcfce7", Verdict.TARGET: "#fef9c3", Verdict.OUT: "#f1f5f9"}
_VERDICT_FG = {Verdict.PRIORITY: "#166534", Verdict.TARGET: "#854d0e", Verdict.OUT: "#475569"}


_EMOJI = re.compile("[\U0001F000-\U0001FAFF\U00002600-\U000027BF\U0001F1E6-\U0001F1FF\uFE0F]+")


def _strip_emoji(text: str) -> str:
    return re.sub(r"\s{2,}", " ", _EMOJI.sub("", text)).strip()


def _e(v) -> str:
    return html.escape(str(v)) if v not in (None, "") else "—"


def build_pdf_html(prospects: list[Prospect], title: str, subtitle: str = "") -> str:
    """HTML compact (sous-ensemble supporté par QTextDocument) : synthèse, tableau, puis fiches."""
    title = _strip_emoji(title)
    now = dt.datetime.now().strftime("%d/%m/%Y %H:%M")
    n_pri = sum(1 for p in prospects if p.verdict is Verdict.PRIORITY)
    n_tar = sum(1 for p in prospects if p.verdict is Verdict.TARGET)
    parts = [f"""
    <h1 style="color:#4c3fd6; font-size:20pt; margin:0">Prospexia — {_e(title)}</h1>
    <p style="color:#64748b; margin:2px 0 10px 0">{_e(subtitle)}<br/>Généré le {now} · {len(prospects)} prospect(s)
       · <b style="color:#166534">{n_pri} prioritaire(s)</b> · <b style="color:#854d0e">{n_tar} cible(s)</b></p>
    <table width="100%" cellspacing="0" cellpadding="4" style="font-size:8.5pt">
      <tr bgcolor="#4c3fd6">
        <th align="left" style="color:white">Entreprise</th><th align="left" style="color:white">Ville</th>
        <th align="left" style="color:white">Téléphone</th><th align="left" style="color:white">Site</th>
        <th align="left" style="color:white">Verdict</th><th align="left" style="color:white">Opportunité</th>
        <th align="left" style="color:white">Identifiant</th>
      </tr>"""]
    for i, p in enumerate(prospects):
        bg = "#ffffff" if i % 2 == 0 else "#f8fafc"
        ident = (f"{p.legal.identifier_label} {p.legal.identifier}".strip() if p.legal and p.legal.identifier else "—")
        parts.append(f"""
      <tr bgcolor="{bg}">
        <td><b>{_e(p.company.name)}</b><br/><span style="color:#64748b">{_e(p.company.sector)}</span></td>
        <td>{_e(p.company.city)}</td><td>{_e(p.company.phone)}</td>
        <td>{_e(p.website.status.label)}</td>
        <td bgcolor="{_VERDICT_BG[p.verdict]}" style="color:{_VERDICT_FG[p.verdict]}"><b>{_e(p.verdict.label)}</b> ({p.score})</td>
        <td>{_e(p.opportunity)}</td><td>{_e(ident)}</td>
      </tr>""")
    parts.append("</table>")

    parts.append('<h2 style="color:#4c3fd6; font-size:14pt; margin-top:18px">Fiches détaillées</h2>')
    for p in prospects:
        c, lg = p.company, p.legal
        rows = [
            ("Secteur", c.sector), ("Adresse", c.address), ("Téléphone", c.phone),
            ("Site web", c.website or "Aucun"), ("État du site", p.website.status.label + (
                " — " + ", ".join(p.website.issues[:3]) if p.website.issues else "")),
            ("Réseaux sociaux", ", ".join(f"{k} : {v}" for k, v in p.presence.socials.items()) or "Aucun trouvé"),
            ("Avis Google", f"{c.reviews_count}" + (f" (note {c.rating})" if c.rating else "")),
            ("Score présence", f"{p.score} / 100 (0 = invisible)"),
            ("Opportunité", p.opportunity),
            ("Diagnostic", " · ".join(p.reasons)),
        ]
        if lg and lg.matched:
            rows += [
                ("Registre", lg.registry),
                (lg.identifier_label or "Identifiant", lg.identifier),
                (lg.secondary_label or "Identifiant 2", lg.secondary_id),
                ("Dénomination légale", lg.legal_name), ("Forme juridique", lg.legal_form),
                ("Activité", f"{lg.activity_code} {lg.activity_label}".strip()),
                ("Création", lg.creation_date), ("Effectif", lg.headcount), ("État", lg.status),
                ("Siège", lg.address), ("Dirigeants", " ; ".join(lg.managers)),
                ("N° TVA", lg.vat_number), ("Fiche officielle", lg.source_url),
            ]
        elif lg:
            rows += [("Registre", f"{lg.registry} — aucune fiche appariée automatiquement"),
                     ("Rechercher", lg.source_url)]
        if c.maps_url:
            rows.append(("Fiche carte", c.maps_url))
        parts.append(f"""
        <table width="100%" cellspacing="0" cellpadding="3" style="font-size:8.5pt; margin-top:8px">
          <tr bgcolor="{_VERDICT_BG[p.verdict]}"><td colspan="2" style="color:{_VERDICT_FG[p.verdict]}">
            <b style="font-size:11pt">{_e(c.name)}</b> &nbsp; {_e(p.verdict.label)} · {_e(c.city)}</td></tr>""")
        for k, v in rows:
            if v in ("", None):
                continue
            parts.append(f'<tr><td width="20%" style="color:#64748b"><b>{_e(k)}</b></td><td>{_e(v)}</td></tr>')
        parts.append("</table>")
    return "\n".join(parts)


def export_pdf(prospects: list[Prospect], path: Path, title: str, subtitle: str = "") -> Path:
    """Génère le PDF via le moteur de rendu Qt (nécessite une QGuiApplication active)."""
    from PySide6.QtCore import QMarginsF
    from PySide6.QtGui import QPageLayout, QPageSize, QPdfWriter, QTextDocument

    path.parent.mkdir(parents=True, exist_ok=True)
    writer = QPdfWriter(str(path))
    writer.setPageSize(QPageSize(QPageSize.PageSizeId.A4))
    writer.setPageOrientation(QPageLayout.Orientation.Portrait)
    writer.setPageMargins(QMarginsF(12, 12, 12, 12), QPageLayout.Unit.Millimeter)
    writer.setTitle(f"Prospexia — {_strip_emoji(title)}")
    writer.setResolution(96)
    doc = QTextDocument()
    doc.setDefaultStyleSheet("body { font-family: 'DejaVu Sans', 'Noto Sans', sans-serif; font-size: 9pt; color: #0f172a; }")
    doc.setHtml(f"<body>{build_pdf_html(prospects, title, subtitle)}</body>")
    doc.setPageSize(writer.pageLayout().paintRectPixels(writer.resolution()).size().toSizeF())
    doc.print_(writer)
    return path
